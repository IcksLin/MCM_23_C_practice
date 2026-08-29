# -*- coding: utf-8 -*-
"""Atomic and structure-preserving ``result2.xlsx`` OOXML export."""
from __future__ import annotations

import hashlib
import os
import posixpath
import re
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

from .preprocess import ModelData


NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", NS)
ET.register_namespace("r", REL_NS)


def _col_letter_to_num(ref: str) -> int:
    match = re.match(r"([A-Z]+)\d+", ref or "")
    if not match:
        return 0
    value = 0
    for char in match.group(1):
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def _col_num_to_letter(col: int) -> str:
    letters = []
    while col:
        col, rem = divmod(col - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def _patch_sheet_xml(xml_bytes: bytes, cell_values: dict) -> bytes:
    """Set numeric values while preserving existing cell style attributes."""
    root = ET.fromstring(xml_bytes)
    sheet_data = root.find(f"{{{NS}}}sheetData")
    if sheet_data is None:
        raise ValueError("worksheet has no sheetData")
    rows = {int(row.get("r")): row for row in sheet_data.findall(f"{{{NS}}}row")}
    for (row_no, col_no), value in sorted(cell_values.items()):
        row = rows.get(row_no)
        if row is None:
            row = ET.Element(f"{{{NS}}}row", {"r": str(row_no)})
            sheet_data.append(row)
            rows[row_no] = row
        ref = f"{_col_num_to_letter(col_no)}{row_no}"
        cell = next((c for c in row.findall(f"{{{NS}}}c") if c.get("r") == ref), None)
        if cell is None:
            cell = ET.SubElement(row, f"{{{NS}}}c", {"r": ref})
        cell.set("t", "n")
        for child_name in ("f", "is"):
            child = cell.find(f"{{{NS}}}{child_name}")
            if child is not None:
                cell.remove(child)
        value_node = cell.find(f"{{{NS}}}v")
        if value_node is None:
            value_node = ET.SubElement(cell, f"{{{NS}}}v")
        value_node.text = format(float(value), ".15g")
        row[:] = sorted(row, key=lambda c: _col_letter_to_num(c.get("r", "")))
    sheet_data[:] = sorted(sheet_data, key=lambda r: int(r.get("r", "0")))
    return ET.tostring(root, encoding="UTF-8", xml_declaration=True)


def _build_cell_values(plan: dict, data: ModelData) -> dict[str, dict]:
    values = {str(year): {} for year in data.years}
    for (j, i, year, season), area in plan["x"].items():
        plot_name = data.plot_names[j]
        col = data.tpl_crop_col.get(i)
        row = (data.tpl_row_s1.get(plot_name) if season == 1
               else data.tpl_row_s2.get(plot_name))
        if col is None or row is None or str(year) not in values:
            if abs(area) > 1e-8:
                raise ValueError(f"plan cell cannot be mapped: {(j, i, year, season)}")
            continue
        key = (row, col)
        values[str(year)][key] = values[str(year)].get(key, 0.0) + float(area)
    return values


def _sheet_paths(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    targets = {rel.get("Id"): rel.get("Target")
               for rel in rels.findall(f"{{{PKG_REL_NS}}}Relationship")}
    result = {}
    sheets = workbook.find(f"{{{NS}}}sheets")
    if sheets is None:
        return result
    for sheet in sheets.findall(f"{{{NS}}}sheet"):
        target = targets.get(sheet.get(f"{{{REL_NS}}}id"))
        if target:
            target = target.lstrip("/")
            if not target.startswith("xl/"):
                target = posixpath.normpath(posixpath.join("xl", target))
            result[sheet.get("name")] = target
    return result


def _sha(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _without_sheet_data(xml_bytes: bytes) -> bytes:
    root = ET.fromstring(xml_bytes)
    sheet_data = root.find(f"{{{NS}}}sheetData")
    if sheet_data is not None:
        root.remove(sheet_data)
    return ET.tostring(root, encoding="UTF-8")


def export_result2_workbook(plan: dict, data: ModelData,
                            template_path: Path, output_path: Path) -> dict:
    """生成、审计并原子发布问题2结果工作簿。"""
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    expected = _build_cell_values(plan, data)
    patch_values = {year: {cell: value for cell, value in cells.items()
                           if abs(value) > 1e-10}
                    for year, cells in expected.items()}
    handle = tempfile.NamedTemporaryFile(prefix=".result2-", suffix=".xlsx",
                                         dir=output_path.parent, delete=False)
    temp_path = Path(handle.name)
    handle.close()
    try:
        with zipfile.ZipFile(template_path, "r") as source:
            sheet_paths = _sheet_paths(source)
            target_sheets = {sheet_paths[str(year)] for year in data.years
                             if str(year) in sheet_paths}
            if len(target_sheets) != len(data.years):
                raise ValueError("template does not contain all seven year sheets")
            with zipfile.ZipFile(temp_path, "w") as target:
                for item in source.infolist():
                    payload = source.read(item.filename)
                    for year, sheet_path in sheet_paths.items():
                        if item.filename == sheet_path and year in patch_values:
                            payload = _patch_sheet_xml(payload, patch_values[year])
                            break
                    target.writestr(item, payload)

        non_target_diff = structure_diff = changed_sheets = 0
        with zipfile.ZipFile(template_path, "r") as original, \
             zipfile.ZipFile(temp_path, "r") as candidate:
            if set(original.namelist()) != set(candidate.namelist()):
                raise ValueError("OOXML member list changed")
            for name in original.namelist():
                before, after = original.read(name), candidate.read(name)
                if name in target_sheets:
                    changed_sheets += int(_sha(before) != _sha(after))
                    structure_diff += int(
                        _without_sheet_data(before) != _without_sheet_data(after))
                else:
                    non_target_diff += int(_sha(before) != _sha(after))

        workbook = openpyxl.load_workbook(temp_path, data_only=True, read_only=True)
        max_diff = 0.0
        nonzero_count = 0
        for year, cells in expected.items():
            if year not in workbook.sheetnames:
                raise ValueError(f"missing output sheet {year}")
            sheet = workbook[year]
            for (row, col), expected_value in cells.items():
                actual = sheet.cell(row, col).value
                actual_value = 0.0 if actual in (None, "") else float(actual)
                max_diff = max(max_diff, abs(actual_value - expected_value))
                nonzero_count += int(abs(actual_value) > 1e-10)
        workbook.close()
        audit = {
            "max_roundtrip_diff": max_diff,
            "non_target_diff_count": non_target_diff,
            "structure_diff_count": structure_diff,
            "changed_sheet_count": changed_sheets,
            "nonzero_cell_count": nonzero_count,
        }
        if (max_diff > 1e-4 or non_target_diff or structure_diff
                or changed_sheets != len(data.years) or nonzero_count == 0):
            raise RuntimeError(f"result2 OOXML audit failed: {audit}")
        os.replace(temp_path, output_path)
        return audit
    finally:
        temp_path.unlink(missing_ok=True)
