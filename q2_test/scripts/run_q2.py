# -*- coding: utf-8 -*-
"""Q2 主入口 — 情景随机 MILP 求解流水线

功能：输入校验 → 清洗 → LHS情景生成 → PAM缩减 → 风险前沿 → 唯一方案选择
      → 样本外评估 → Excel回填 → 图表 → 报告

运行环境：Anaconda/Miniconda Python 3.10+, 已安装 requirements.txt 依赖

使用方法：
    cd q2_test
    python scripts/run_q2.py --seed 2024 --raw-scenarios 1000 --reduced-scenarios 30 --beta 0.90 --lambda-grid 0:1:0.1 --out-sample 5000 --mip-gap 0.001 --time-limit 600

冒烟测试（快速验证）：
    python scripts/run_q2.py --raw-scenarios 20 --reduced-scenarios 5 --out-sample 50 --time-limit 30 --mip-gap 0.05 --figures
"""
from __future__ import annotations
import sys
import os
import argparse
import hashlib
import importlib.metadata
import json
import platform
import time
from pathlib import Path
import numpy as np
import pandas as pd

# 将 algorithms 加入 import 路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from algorithms import paths, io_data, preprocess
from algorithms.scenarios import generate_raw_scenarios, compute_proxy_profit
from algorithms.scenario_reduction import reduce_scenarios
from algorithms.solve import solve_lexicographic, extract_solution
from algorithms.risk import (recompute_scenario_profits, compute_cvar,
                             select_unique_plan, pareto_nondominated)
from algorithms.validate import validate_solution
from algorithms.export_ooxml import patch_result2
from algorithms.plots import generate_figures


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def _print_progress(pct, msg):
    """打印文本进度条。"""
    bar_len = 30
    filled = int(pct * bar_len / 100)
    bar = "#" * filled + "." * (bar_len - filled)
    print(f"\r[{bar}] {pct:3.0f}%  {msg}", end="", flush=True)
    if pct >= 100:
        print()


def _load_q1_baseline(data):
    """从Q1 result1_1.xlsx读取基线方案，用于情景缩减的代理利润计算。

    如果文件不存在，回退到2023实际种植方案。
    """
    plan_x = {}
    if paths.Q1_RESULT1_1.exists():
        import openpyxl
        wb = openpyxl.load_workbook(paths.Q1_RESULT1_1, data_only=True)
        crop_by_col = {col: crop for crop, col in data.tpl_crop_col.items()}
        for year in data.years:
            if str(year) not in wb.sheetnames:
                continue
            ws = wb[str(year)]
            for plot_name, row in data.tpl_row_s1.items():
                j = data.plot_idx.get(plot_name)
                if j is None:
                    continue
                for col, crop in crop_by_col.items():
                    val = ws.cell(row, col).value
                    if val not in (None, "") and float(val) > 0:
                        plan_x[(j, crop, year, 1)] = float(val)
            for plot_name, row in data.tpl_row_s2.items():
                j = data.plot_idx.get(plot_name)
                if j is None:
                    continue
                for col, crop in crop_by_col.items():
                    val = ws.cell(row, col).value
                    if val not in (None, "") and float(val) > 0:
                        plan_x[(j, crop, year, 2)] = float(val)
        wb.close()
    else:
        # 回退：使用2023实际种植作为代理基线
        for (j, i, s), area in data.bar_x.items():
            for t in data.years:
                plan_x[(j, i, t, s)] = area
    return plan_x


def _generate_report(data, frontier, selected_lambda, evaluation,
                      audit, solve_info, args):
    """生成文字报告。"""
    report_path = paths.DOC_DIR / "Q2_建模实现报告.md"
    report_path.parent.mkdir(parents=True, exist_ok=True)

    lines = [
        "# 问题 2 建模实现报告",
        "",
        f"> 生成时间: {time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"> 随机种子: {args.seed}",
        f"> 原始情景数: {args.raw_scenarios}",
        f"> 缩减情景数: {args.reduced_scenarios}",
        f"> CVaR置信水平: {args.beta}",
        f"> 样本外情景数: {args.out_sample}",
        "",
        "## 1. 数据与断言",
        "",
        f"- 地块数: {data.assertions['n_plots']}",
        f"- 作物数: {data.assertions['n_crops']}",
        f"- 2023种植记录: {data.assertions['n_planting_2023']}",
        f"- 总面积: {data.assertions['total_area']:.0f} 亩",
        f"- 智慧大棚继承产量: {data.assertions['smart_s1_inherited_yield']:.0f} 斤",
        f"- 智慧大棚继承成本: {data.assertions['smart_s1_inherited_cost']:.0f} 元",
        "",
        "## 2. 模型概要",
        "",
        "- 共同决策变量: x[j,i,t,s] (种植面积), y[j,i,t,s] (0-1激活), r[j,t] (水浇地模式)",
        "- 情景变量: Q_omega (产量), u_omega (销量), xi_omega (CVaR超额), zeta (VaR代理)",
        "- 目标: max Z_lambda = (1-lambda)*E[Pi] + lambda*LCVaR_beta",
        "- 约束: 继承Q1全部农业硬约束 + 情景产量/销量/CVaR约束",
        "- 求解: 三级字典序 (风险 → 期望利润 → 最小化激活)",
        "",
        "## 3. 风险前沿",
        "",
        "| lambda | 期望利润 | 下尾CVaR | 激活数 | 状态 | gap |",
        "|---|---|---|---|---|---|",
    ]

    if frontier is not None:
        for _, row in frontier.iterrows():
            lines.append(
                f"| {row['lambda']} | {row['expected_profit']:.0f} | "
                f"{row['lower_tail_cvar']:.0f} | {row.get('n_activations', '-')} | "
                f"{row.get('status', '?')} | {row.get('mip_gap', 0):.4f} |"
            )

    nondominated = pareto_nondominated(
        frontier[frontier.get("eligible", True)]) if frontier is not None else pd.DataFrame()
    if len(nondominated) == 1:
        selection_reason = "唯一非支配可用方案"
    else:
        selection_reason = "非支配风险前沿的确定性膝点规则"

    lines.extend([
        "",
        f"## 4. 唯一方案选择",
        "",
        f"- 选择 lambda = {selected_lambda}",
        f"- 选择理由: {selection_reason}",
        f"- 求解认证: {'通过' if audit.get('certified', False) else '未认证（保留可行解）'}",
        "",
        "## 5. 样本外评估",
        "",
    ])

    if evaluation:
        lines.append(f"- 平均利润: {evaluation.get('mean_profit', 0):.0f} 元")
        lines.append(f"- 标准差: {evaluation.get('std_profit', 0):.0f} 元")
        lines.append(f"- 10%分位数: {evaluation.get('p10_profit', 0):.0f} 元")
        lines.append(f"- 最低利润: {evaluation.get('min_profit', 0):.0f} 元")
        lines.append(f"- 亏损概率: {evaluation.get('loss_prob', 0):.4f}")

    lines.extend([
        "",
        "## 6. 约束审计",
        "",
        f"- 最大面积守恒违约: {audit.get('max_area_violation', 0):.2e}",
        f"- 最大不适种面积: {audit.get('max_unsuitable_area', 0):.2e}",
        f"- 重茬违反次数: {audit.get('rotation_violation_count', 0)}",
        f"- 滚动三年豆类最小裕度: {audit.get('min_legume_slack', 0):.2f}",
        f"- 水浇地模式冲突: {audit.get('irrigated_mode_conflicts', 0)}",
        f"- 0-1整数性违约: {audit.get('max_integrality_violation', 0):.2e}",
        f"- Excel回读差: {audit.get('excel_roundtrip_diff', 0):.2e}",
        f"- 可行性: {'通过' if audit.get('feasible', False) else '未通过'}",
        "",
        "## 7. 交付物清单",
        "",
        f"- 结果工作簿: outputs/q2/result2.xlsx",
        f"- 可读种植方案: outputs/q2/selected_plan.csv",
        f"- 情景摘要: outputs/q2/scenario_summary.csv",
        f"- 风险前沿: outputs/q2/risk_frontier.csv",
        f"- 样本外利润: outputs/q2/out_of_sample_profits.csv",
        f"- 样本外指标: outputs/q2/out_of_sample_metrics.csv",
        f"- 约束审计: outputs/q2/audit_q2.csv",
        f"- 复现清单: outputs/q2/repro_q2.json",
        f"- 图表: outputs/q2/figures/",
        f"- 本报告: doc/Q2_建模实现报告.md",
        "",
        "## 8. 复现命令",
        "",
        "```powershell",
        f"python scripts/run_q2.py --seed {args.seed} --raw-scenarios {args.raw_scenarios} "
        f"--reduced-scenarios {args.reduced_scenarios} --beta {args.beta} "
        f"--lambda-grid {args.lambda_grid} --out-sample {args.out_sample} "
        f"--mip-gap {args.mip_gap} --time-limit {args.time_limit}",
        "```",
    ])

    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path


def main():
    parser = argparse.ArgumentParser(description="2024 C题 问题2 求解主程序")
    parser.add_argument("--config", type=str, default=None,
                        help="YAML参数文件（命令行显式参数优先）")
    parser.add_argument("--seed", type=int, default=2024, help="随机种子 (默认2024)")
    parser.add_argument("--raw-scenarios", type=int, default=1000, help="原始情景数 (默认1000)")
    parser.add_argument("--reduced-scenarios", type=int, default=30, help="缩减情景数 (默认30)")
    parser.add_argument("--beta", type=float, default=0.90, help="CVaR置信水平 (默认0.90)")
    parser.add_argument("--lambda-grid", type=str, default="0:1:0.1",
                        help="lambda网格 start:stop:step (默认 0:1:0.1)")
    parser.add_argument("--out-sample", type=int, default=5000, help="样本外情景数 (默认5000)")
    parser.add_argument("--mip-gap", type=float, default=0.001, help="MIP最优性间隙 (默认0.001)")
    parser.add_argument("--time-limit", type=float, default=600.0, help="求解时间限制(秒) (默认600)")
    parser.add_argument("--eta", type=float, default=0.5, help="最小种植面积比例 (默认0.5)")
    parser.add_argument("--figures", action="store_true", help="生成图表")
    parser.add_argument("--reports", action="store_true", help="生成文字报告")
    parser.add_argument("--distribution", choices=["uniform", "triangular"], default="uniform",
                        help="分布假设 (默认uniform)")
    parser.add_argument("--allow-uncertified", action="store_true",
                        help="仅供P1/冒烟测试：允许未认证可行解以退出码0结束")
    args = parser.parse_args()
    if args.config:
        import yaml
        cfg = yaml.safe_load(Path(args.config).read_text(encoding="utf-8")) or {}
        provided = set(sys.argv[1:])
        mapping = {
            "seed": "--seed", "raw_scenarios": "--raw-scenarios",
            "reduced_scenarios": "--reduced-scenarios", "beta": "--beta",
            "lambda_grid": "--lambda-grid", "out_sample": "--out-sample",
            "mip_gap": "--mip-gap", "time_limit": "--time-limit",
            "eta": "--eta", "figures": "--figures", "reports": "--reports",
            "distribution": "--distribution",
            "allow_uncertified": "--allow-uncertified",
        }
        for key, flag in mapping.items():
            if key in cfg and flag not in provided:
                setattr(args, key, cfg[key])

    paths.ensure_dirs()
    start_time = time.time()

    # ---- 1. 输入校验 ----
    _print_progress(5, "输入校验")
    ver = io_data.verify_inputs()
    bad = [p for p, (_, ok, _) in ver.items() if not ok]
    if bad:
        print(f"\n[ERROR] 输入哈希不匹配: {bad}")
        sys.exit(1)

    # ---- 2. 数据加载与清洗 ----
    _print_progress(10, "数据清洗")
    raw = io_data.load_inputs()
    data = preprocess.preprocess(raw)

    # ---- 3. 情景生成 ----
    _print_progress(20, f"LHS情景生成 (N={args.raw_scenarios})")
    scenarios = generate_raw_scenarios(
        data, n=args.raw_scenarios, seed=args.seed,
        distribution=args.distribution)
    # 导出情景摘要
    scenario_summary = pd.DataFrame([{
        "n_scenarios": scenarios.n,
        "seed": scenarios.seed,
        "distribution": scenarios.distribution,
        "n_demand_params": len(scenarios.demand),
        "n_yield_params": len(scenarios.yield_),
        "n_price_params": len(scenarios.price),
    }])
    scenario_summary.to_csv(paths.SCENARIO_SUMMARY, index=False)

    # ---- 4. 情景缩减 ----
    _print_progress(30, f"PAM缩减 (K={args.reduced_scenarios})")
    baseline_plan = _load_q1_baseline(data)
    reduced = reduce_scenarios(data, scenarios, k=args.reduced_scenarios,
                               baseline_plan=baseline_plan, seed=args.seed)

    # ---- 5. 风险前沿 ----
    # parse lambda grid
    parts = args.lambda_grid.split(":")
    lam_start = float(parts[0]); lam_stop = float(parts[1]); lam_step = float(parts[2])
    lambdas = list(np.arange(lam_start, lam_stop + lam_step / 2, lam_step))
    lambdas = [round(x, 4) for x in lambdas]

    frontier_rows = []
    results_by_lambda = {}

    pct_start, pct_end = 30, 75
    for li, lam in enumerate(lambdas):
        pct = int(pct_start + (pct_end - pct_start) * li / max(1, len(lambdas) - 1))
        _print_progress(pct, f"lambda={lam:.1f} 三级字典序求解")

        lex = solve_lexicographic(
            data, reduced, beta=args.beta, risk_lambda=lam,
            eta=args.eta, time_limit=args.time_limit,
            mip_gap=args.mip_gap, eps_z=None, eps_e=None)

        if lex["z_star"] is not None and lex["solution"] is not None:
            sol = lex["solution"]
            # 独立复算利润
            profits = recompute_scenario_profits(sol["x"], reduced, data)
            e_pi = float(np.average(profits, weights=reduced.weights))
            cvar = compute_cvar(profits, reduced.weights, args.beta)
            z_eval = (1.0 - lam) * e_pi + lam * cvar
            stage_results = [lex.get(f"result{k}") for k in (1, 2, 3)]
            stage_results = [res for res in stage_results if res is not None]
            certified = bool(
                lex.get("lex_complete", False)
                and stage_results
                and all(res.status == 0 and np.isfinite(res.fun)
                        and (not np.isfinite(res.mip_gap)
                             or res.mip_gap <= args.mip_gap + 1e-12)
                        for res in stage_results)
            )
            frontier_rows.append({
                "lambda": lam,
                "z_lambda": z_eval,
                "expected_profit": e_pi,
                "lower_tail_cvar": cvar,
                "n_activations": sol["n_activations"],
                "status": lex["result"].solver_status,
                "mip_gap": lex["result"].mip_gap,
                "stage1_gap": lex.get("result1", lex["result"]).mip_gap,
                "stage2_gap": lex.get("result2").mip_gap if lex.get("result2") else float("nan"),
                "stage3_gap": lex.get("result3").mip_gap if lex.get("result3") else float("nan"),
                "dual_bound": lex["result"].dual_bound,
                "solve_time": sum(res.time for res in stage_results),
                "final_stage": lex.get("final_stage", 1),
                "lex_complete": bool(lex.get("lex_complete", False)),
                "certified": certified,
                "eligible": bool(lex.get("final_stage", 1) >= 2),
            })
            if lex.get("final_stage", 1) >= 2:
                results_by_lambda[str(lam)] = lex
        else:
            frontier_rows.append({
                "lambda": lam,
                "z_lambda": np.nan,
                "expected_profit": np.nan,
                "lower_tail_cvar": np.nan,
                "n_activations": 0,
                "status": getattr(lex.get("result1", lex.get("result")), "solver_status", "unknown"),
                "mip_gap": 1.0,
                "stage1_gap": getattr(lex.get("result1"), "mip_gap", float("nan")) if lex.get("result1") else float("nan"),
                "stage2_gap": float("nan"),
                "stage3_gap": float("nan"),
                "dual_bound": np.nan,
                "solve_time": np.nan,
                "final_stage": 0,
                "lex_complete": False,
                "certified": False,
                "eligible": False,
            })

    frontier_df = pd.DataFrame(frontier_rows)
    frontier_df.to_csv(paths.RISK_FRONTIER_CSV, index=False)

    # ---- 6. 唯一方案选择 ----
    _print_progress(77, "风险前沿膝点选择")
    feasible_frontier = frontier_df[frontier_df["eligible"]].reset_index(drop=True)
    frontier_complete = len(feasible_frontier) == len(lambdas)
    if len(feasible_frontier) > 0:
        selected_lambda = select_unique_plan(feasible_frontier)
    else:
        selected_lambda = None

    if selected_lambda and selected_lambda in results_by_lambda:
        lex = results_by_lambda[selected_lambda]
        plan = lex["solution"]
    else:
        print("\n[WARNING] 无可行方案，无法继续")
        sys.exit(1)

    # ---- 7. 样本外评估 ----
    _print_progress(85, f"样本外评估 (N={args.out_sample})")
    oos_scenarios = generate_raw_scenarios(
        data, n=args.out_sample, seed=args.seed + 99999,
        distribution=args.distribution)
    oos_profits = recompute_scenario_profits(plan["x"], oos_scenarios, data)
    comparison = {"selected": oos_profits}
    if baseline_plan:
        comparison["q1_baseline"] = recompute_scenario_profits(
            baseline_plan, oos_scenarios, data)
    neutral = results_by_lambda.get("0.0")
    if neutral is not None:
        comparison["risk_neutral"] = recompute_scenario_profits(
            neutral["solution"]["x"], oos_scenarios, data)
    oos_df = pd.DataFrame(comparison)
    oos_df.to_csv(paths.OUT_OF_SAMPLE_PROFITS, index=False)

    oos_metrics = {
        "mean_profit": float(oos_profits.mean()),
        "std_profit": float(oos_profits.std()),
        "p10_profit": float(np.percentile(oos_profits, 10)),
        "min_profit": float(oos_profits.min()),
        "loss_prob": float(np.mean(oos_profits < 0)),
        "worst_10pct_avg": float(np.mean(np.sort(oos_profits)[:max(1, len(oos_profits)//10)])),
    }
    pd.DataFrame([oos_metrics]).to_csv(paths.OUT_OF_SAMPLE_METRICS, index=False)

    # ---- 8. Independent audit before publishing Excel ----
    selected_profits = recompute_scenario_profits(plan["x"], reduced, data)
    selected_e = float(np.average(selected_profits, weights=reduced.weights))
    selected_cvar = compute_cvar(selected_profits, reduced.weights, args.beta)
    selected_z = ((1.0 - float(selected_lambda)) * selected_e
                  + float(selected_lambda) * selected_cvar)
    selected_row = feasible_frontier[
        feasible_frontier["lambda"].astype(str) == selected_lambda].iloc[0]
    solver_info = {
        "status": lex["result"].status,
        "dual_bound": lex["result"].dual_bound,
        "mip_gap": lex["result"].mip_gap,
        "time": lex["result"].time,
        "certified": bool(selected_row["certified"]),
    }
    pre_audit = validate_solution(
        plan, data, reduced, beta=args.beta,
        risk_lambda=float(selected_lambda), z_lambda=selected_z,
        e_pi=selected_e, cvar_value=selected_cvar, eta=args.eta,
        solver_info=solver_info)
    if not pre_audit["feasible"]:
        pd.DataFrame([pre_audit]).to_csv(paths.AUDIT_PATH, index=False)
        print("\n[ERROR] 内存方案约束审计未通过，禁止写入 result2.xlsx")
        sys.exit(1)

    # ---- 9. Excel patch + post-export audit ----
    _print_progress(92, "Excel模板回填与结构审计")
    excel_audit = patch_result2(
        plan, data, paths.TEMPLATE2_PATH, paths.RESULT2_PATH)
    _print_progress(95, "最终约束审计")
    audit = validate_solution(
        plan, data, reduced, beta=args.beta,
        risk_lambda=float(selected_lambda), z_lambda=selected_z,
        e_pi=selected_e, cvar_value=selected_cvar, eta=args.eta,
        excel_audit=excel_audit, solver_info=solver_info)
    pd.DataFrame([audit]).to_csv(paths.AUDIT_PATH, index=False)

    plan_rows = [{"plot": data.plot_names[j], "crop_code": i, "year": t,
                  "season": s, "area": area}
                 for (j, i, t, s), area in plan["x"].items() if area > 1e-8]
    pd.DataFrame(plan_rows).to_csv(paths.SELECTED_PLAN_CSV, index=False)

    # ---- 10. 图表 ----
    if args.figures:
        _print_progress(98, "生成图表")
        generate_figures(data, scenarios, reduced,
                         frontier_df, comparison, plan, comparison)

    # ---- 11. 报告 ----
    if args.reports:
        _print_progress(99, "生成报告")
        solve_info = {"lex": lex}
        _generate_report(data, frontier_df, selected_lambda,
                         oos_metrics, audit, solve_info, args)

    # ---- 12. 复现清单 ----
    repro = {
        "seed": args.seed,
        "raw_scenarios": args.raw_scenarios,
        "reduced_scenarios": args.reduced_scenarios,
        "beta": args.beta,
        "lambda_grid": args.lambda_grid,
        "out_sample": args.out_sample,
        "mip_gap": args.mip_gap,
        "time_limit": args.time_limit,
        "distribution": args.distribution,
        "selected_lambda": selected_lambda,
        "frontier_complete": frontier_complete,
        "selected_certified": bool(selected_row["certified"]),
        "f1_sha": data.f1_sha,
        "f2_sha": data.f2_sha,
        "template2_sha": data.template2_sha,
        "elapsed_seconds": time.time() - start_time,
        "versions": {
            "python": platform.python_version(),
            **{name: importlib.metadata.version(name)
               for name in ("numpy", "pandas", "scipy", "openpyxl", "matplotlib")},
            "solver": "HiGHS via scipy.optimize.milp",
        },
        "command": "python scripts/run_q2.py " + " ".join(sys.argv[1:]),
    }
    output_hashes = {}
    report_file = paths.DOC_DIR / "Q2_建模实现报告.md"
    outputs_to_hash = [p for p in paths.Q2_OUT_DIR.rglob("*")
                       if p.is_file() and p != paths.REPRO_PATH]
    if report_file.exists():
        outputs_to_hash.append(report_file)
    for output in outputs_to_hash:
        if output.exists():
            key = (str(output.relative_to(paths.Q2_ROOT))
                   if output.is_relative_to(paths.Q2_ROOT) else str(output))
            output_hashes[key] = hashlib.sha256(output.read_bytes()).hexdigest()
    repro["output_sha256"] = output_hashes
    paths.REPRO_PATH.write_text(
        json.dumps(repro, indent=2, ensure_ascii=False), encoding="utf-8")

    # ---- 最终摘要 ----
    _print_progress(100, "完成")
    print()
    print("=" * 60)
    print("              Q2 求解完成 — 最终摘要")
    print("=" * 60)
    print(f"选择 lambda:     {selected_lambda}")
    print(f"期望利润:       {oos_metrics['mean_profit']:.0f} 元")
    print(f"10%利润分位:    {oos_metrics['p10_profit']:.0f} 元")
    print(f"最差10%均值:    {oos_metrics['worst_10pct_avg']:.0f} 元")
    print(f"最低利润:       {oos_metrics['min_profit']:.0f} 元")
    print(f"亏损概率:       {oos_metrics['loss_prob']:.4f}")
    print(f"约束可行:       {'通过' if audit.get('feasible') else '未通过'}")
    print(f"Excel回读差:    {excel_audit['max_roundtrip_diff']:.2e}")
    print(f"风险前沿完整:   {'是' if frontier_complete else '否'}")
    print(f"求解结果认证:   {'是' if selected_row['certified'] else '否'}")
    print(f"总耗时:         {time.time() - start_time:.1f} 秒")
    print("=" * 60)
    delivery_ready = bool(audit.get("feasible") and frontier_complete
                          and selected_row["certified"])
    if not delivery_ready and not args.allow_uncertified:
        print("[WARNING] 已保存可行候选解，但未达到正式交付门槛。")
        sys.exit(2)


if __name__ == "__main__":
    main()
