# -*- coding: utf-8 -*-
"""Read-only Excel ingestion for Q3.

Copied from q2_test/algorithms/io_data.py and extended for Q3 per
AGENT.md section 3.2-3.3.  All Q2 read-only parsers (附件1 plots & crops,
附件2 2023 planting & statistics, result2.xlsx template structure) are
preserved unchanged; the module additionally ingests Q2's selected_plan.csv
baseline and exposes a hash-verification helper.

Q3-specific additions:
  - BaselinePlan dataclass holding Q2's selected_plan.csv as a dict keyed
    by (j, i, t, s) -> area.
  - load_q2_baseline(plan_csv_path) -> BaselinePlan with full column
    validation (columns, types, year/season/crop_code/area ranges).
  - verify_input_hashes(expected) -> None, raising ValueError on any
    SHA-256 mismatch (strict variant of verify_inputs).
"""
from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
import hashlib
import pandas as pd
from openpyxl import load_workbook

from . import paths


def _safe(v):
    if v is None:
        return ""
    return str(v)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest().upper()


def verify_inputs() -> dict:
    """Return {path: (sha, ok, expected)} for every frozen input."""
    out = {}
    for p, exp in paths.EXPECTED_SHA.items():
        sha = sha256_file(p)
        out[str(p)] = (sha, sha == exp, exp)
    return out


def verify_input_hashes(expected: dict) -> None:
    """Strict hash check: raise ValueError if any path's SHA-256 mismatches.

    Args:
        expected: mapping {Path|str: expected_sha_upper_hex}.
    """
    mismatches = []
    for p, exp_sha in expected.items():
        pth = Path(p)
        actual = sha256_file(pth)
        if actual != exp_sha.upper():
            mismatches.append((str(pth), actual, exp_sha))
    if mismatches:
        details = "; ".join(
            f"{p}: got {a}, expected {e}" for p, a, e in mismatches
        )
        raise ValueError(f"Input hash mismatch: {details}")


@dataclass
class RawData:
    plots: pd.DataFrame            # name, type, area
    crops: pd.DataFrame            # code, name, type, lands_raw, note
    planting_2023: pd.DataFrame    # plot, crop_code, crop_name, crop_type, area, season
    stats_2023: pd.DataFrame       # seq, crop_code, crop_name, land_type, season, yield, cost, price_raw
    template_crops: list           # crop names in template column order (len 41)
    template_plot_s1: list         # plot names in 第一季 rows (len 54)
    template_plot_s2: list         # plot names in 第二季 rows (len 28)
    template_years: list           # sheet year labels
    f1_path: Path
    f2_path: Path
    template2_path: Path
    f1_sha: str
    f2_sha: str
    template2_sha: str


@dataclass
class BaselinePlan:
    """Q2 selected_plan.csv parsed into index space.

    Fields:
      plot_idx   : dict[str, int]   plot name -> j index (0-based, by
                                    first-appearance order in the CSV)
      crop_code  : set[int]          distinct crop codes present in the plan
      year       : list[int]        sorted distinct years present in the plan
      season     : list[int]        sorted distinct seasons present in the plan
      area       : dict[tuple, float]  (j, i, t, s) -> planted area (mu)
    """
    plot_idx: dict = field(default_factory=dict)
    crop_code: set = field(default_factory=set)
    year: list = field(default_factory=list)
    season: list = field(default_factory=list)
    area: dict = field(default_factory=dict)


def _is_int_like(v) -> bool:
    if isinstance(v, (int,)) and not isinstance(v, bool):
        return True
    if isinstance(v, float) and v.is_integer():
        return True
    if isinstance(v, str):
        try:
            int(v.strip())
            return True
        except ValueError:
            return False
    return False


def _read_plots(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb["乡村的现有耕地"]
    rows = []
    for r in range(2, ws.max_row + 1):
        name = _safe(ws.cell(r, 1).value)
        typ = _safe(ws.cell(r, 2).value)
        area = ws.cell(r, 3).value
        if name == "" and typ == "":
            continue
        rows.append({"name": name, "type": typ, "area": area})
    wb.close()
    return pd.DataFrame(rows)


def _read_crops(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb["乡村种植的农作物"]
    rows = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not _is_int_like(code):
            continue
        name = _safe(ws.cell(r, 2).value)
        typ = _safe(ws.cell(r, 3).value)
        lands = _safe(ws.cell(r, 4).value)
        note = _safe(ws.cell(r, 5).value)
        rows.append({
            "code": int(code),
            "name": name,
            "type": typ,
            "lands_raw": lands,
            "note": note,
        })
    wb.close()
    return pd.DataFrame(rows)


def _read_planting(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb["2023年的农作物种植情况"]
    rows = []
    prev_block = ""
    for r in range(2, ws.max_row + 1):
        blk = _safe(ws.cell(r, 1).value)
        if blk == "":
            blk = prev_block
        else:
            prev_block = blk
        code = ws.cell(r, 2).value
        name = _safe(ws.cell(r, 3).value)
        typ = _safe(ws.cell(r, 4).value)
        area = ws.cell(r, 5).value
        season = _safe(ws.cell(r, 6).value)
        if code is None and name == "":
            continue
        rows.append({
            "plot": blk,
            "crop_code": code,
            "crop_name": name,
            "crop_type": typ,
            "area": area,
            "season": season,
        })
    wb.close()
    return pd.DataFrame(rows)


def _read_stats(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb["2023年统计的相关数据"]
    rows = []
    for r in range(2, ws.max_row + 1):
        seq = ws.cell(r, 1).value
        code = ws.cell(r, 2).value
        if isinstance(seq, str) and "注" in seq:
            continue
        if not _is_int_like(code):
            continue
        name = _safe(ws.cell(r, 3).value)
        land = _safe(ws.cell(r, 4).value)
        season = _safe(ws.cell(r, 5).value)
        yld = ws.cell(r, 6).value
        cost = ws.cell(r, 7).value
        price = _safe(ws.cell(r, 8).value)
        rows.append({
            "seq": seq,
            "crop_code": int(code),
            "crop_name": name,
            "land_type": land,
            "season": season,
            "yield": yld,
            "cost": cost,
            "price_raw": price,
        })
    wb.close()
    return pd.DataFrame(rows)


def _read_template2(path: Path) -> tuple[list, list, list, list]:
    """Read result2.xlsx structure.

    Returns (crop_names, plot_s1_names, plot_s2_names, year_labels).
    Each year sheet has the same row/column layout as result1.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    year_labels = sorted([ws.title for ws in wb.worksheets], key=int)
    # use first year sheet to get layout
    ws = wb[year_labels[0]]
    crops = []
    for c in range(3, ws.max_column + 1):
        v = _safe(ws.cell(1, c).value)
        if v != "":
            crops.append(v)
    # season-1 plot rows 2..55 (54 plots)
    s1 = [_safe(ws.cell(r, 2).value) for r in range(2, 56)]
    # season-2 plot rows 56..83 (28 plots: 8 water + 16 normal + 4 smart)
    s2 = [_safe(ws.cell(r, 2).value) for r in range(56, 84)]
    wb.close()
    return crops, s1, s2, year_labels


def load_inputs() -> RawData:
    """Read all inputs (read-only). Raises on hash mismatch."""
    ver = verify_inputs()
    bad = [p for p, (_, ok, _) in ver.items() if not ok]
    if bad:
        raise RuntimeError(f"Input hash mismatch, refusing to proceed: {bad}")
    plots = _read_plots(paths.F1_PATH)
    crops = _read_crops(paths.F1_PATH)
    planting = _read_planting(paths.F2_PATH)
    stats = _read_stats(paths.F2_PATH)
    tpl_crops, tpl_s1, tpl_s2, tpl_years = _read_template2(paths.TEMPLATE2_PATH)
    return RawData(
        plots=plots,
        crops=crops,
        planting_2023=planting,
        stats_2023=stats,
        template_crops=tpl_crops,
        template_plot_s1=tpl_s1,
        template_plot_s2=tpl_s2,
        template_years=tpl_years,
        f1_path=paths.F1_PATH,
        f2_path=paths.F2_PATH,
        template2_path=paths.TEMPLATE2_PATH,
        f1_sha=ver[str(paths.F1_PATH)][0],
        f2_sha=ver[str(paths.F2_PATH)][0],
        template2_sha=ver[str(paths.TEMPLATE2_PATH)][0],
    )


# ---- Q2 baseline ingestion (AGENT.md section 5) ----

_REQUIRED_COLS = ("plot", "crop_code", "year", "season", "area")
_VALID_YEARS = range(2024, 2031)        # [2024, 2030]
_VALID_SEASONS = (1, 2)
_CROP_CODE_MIN = 1
_CROP_CODE_MAX = 41


def load_q2_baseline(plan_csv_path) -> BaselinePlan:
    """Read Q2's selected_plan.csv into a BaselinePlan.

    CSV format: plot,crop_code,year,season,area (one row per planting).

    Validates:
      - all required columns exist;
      - crop_code / year / season are int-like;
      - area is numeric;
      - year in [2024, 2030], season in {1, 2},
        crop_code in [1, 41], area >= 0.

    Raises:
        ValueError on any column/type/range violation.
    """
    csv_path = Path(plan_csv_path)
    df = pd.read_csv(csv_path, dtype={"plot": str})

    # ---- column existence ----
    missing_cols = [c for c in _REQUIRED_COLS if c not in df.columns]
    if missing_cols:
        raise ValueError(
            f"selected_plan.csv missing required columns: {missing_cols}"
        )

    # ---- type + range validation per row ----
    bad_rows = []
    for ridx, row in df.iterrows():
        # crop_code: int-like, in [1, 41]
        cc = row["crop_code"]
        if not _is_int_like(cc):
            bad_rows.append((ridx, "crop_code", "not int-like", cc))
        else:
            cc_i = int(cc)
            if not (_CROP_CODE_MIN <= cc_i <= _CROP_CODE_MAX):
                bad_rows.append((ridx, "crop_code", "out of [1,41]", cc_i))
        # year: int-like, in [2024, 2030]
        yr = row["year"]
        if not _is_int_like(yr):
            bad_rows.append((ridx, "year", "not int-like", yr))
        else:
            yr_i = int(yr)
            if yr_i not in _VALID_YEARS:
                bad_rows.append((ridx, "year", "out of [2024,2030]", yr_i))
        # season: int-like, in {1, 2}
        ss = row["season"]
        if not _is_int_like(ss):
            bad_rows.append((ridx, "season", "not int-like", ss))
        else:
            ss_i = int(ss)
            if ss_i not in _VALID_SEASONS:
                bad_rows.append((ridx, "season", "out of {1,2}", ss_i))
        # area: numeric, >= 0
        ar = row["area"]
        try:
            ar_f = float(ar)
        except (TypeError, ValueError):
            bad_rows.append((ridx, "area", "not numeric", ar))
            ar_f = None
        if ar_f is not None and ar_f < 0:
            bad_rows.append((ridx, "area", "negative", ar_f))

    if bad_rows:
        preview = "; ".join(
            f"row {r}({col}): {reason} ({val!r})" for r, col, reason, val in bad_rows[:10]
        )
        raise ValueError(
            f"selected_plan.csv validation failed ({len(bad_rows)} rows): {preview}"
        )

    # ---- build BaselinePlan ----
    bp = BaselinePlan()
    for _, row in df.iterrows():
        plot = _clean_plot_name(row["plot"])
        i = int(row["crop_code"])
        t = int(row["year"])
        s = int(row["season"])
        a = float(row["area"])
        if plot not in bp.plot_idx:
            bp.plot_idx[plot] = len(bp.plot_idx)
        j = bp.plot_idx[plot]
        bp.crop_code.add(i)
        bp.area[(j, i, t, s)] = bp.area.get((j, i, t, s), 0.0) + a

    bp.year = sorted({int(y) for y in df["year"]})
    bp.season = sorted({int(sv) for sv in df["season"]})
    return bp


def _clean_plot_name(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("_x000d_", "").strip()
    return s
