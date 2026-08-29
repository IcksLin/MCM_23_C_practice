# -*- coding: utf-8 -*-
"""Q3 候选 Excel 原子回填 — result3_candidate.xlsx 的 OOXML 直接修改。

功能：
  复制自 q2_test/algorithms/export_ooxml.py 并适配 Q3。Q3 的 plan 结构包含
  b[j,t]（豆类前茬指示）与 w[j,i,t,s]（互补面积线性化变量），但只有 x
  （种植面积）被写入候选 Excel。模板仍是 doc/C题/附件3/result2.xlsx。

  这是候选展示文件（outputs/q3/result3_candidate.xlsx），不是官方要求
  的 result2.xlsx。官方提交文件不在本模块职责范围。

使用方法：
    from algorithms.io_data import load_raw
    from algorithms.preprocess import preprocess
    from algorithms.export_ooxml import patch_result3

    raw = load_raw()
    data = preprocess(raw)
    max_diff = patch_result3(plan, data)   # 默认写到 paths.RESULT3_CANDIDATE
    print(f"回读最大误差: {max_diff}")

命令行示例（已配置 PYTHONPATH 到 q3_test 根目录）：
    python -c "from algorithms.export_ooxml import patch_result3; \
import pickle; plan, data = pickle.load(open('plan.pkl','rb')); \
print(patch_result3(plan, data))"

运行环境：Python 3.10+，依赖 openpyxl（只读校验）与标准库 zipfile/xml。
"""
from __future__ import annotations

import hashlib
import os
import posixpath
import re
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

from .preprocess import ModelData
from . import paths


NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", NS)
ET.register_namespace("r", REL_NS)


def _col_letter_to_num(ref: str) -> int:
    """把单元格引用的列字母（如 'A1'）转成 1 起始的整数列号。"""
    match = re.match(r"([A-Z]+)\d+", ref or "")
    if not match:
        return 0
    value = 0
    for char in match.group(1):
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def _col_num_to_letter(col: int) -> str:
    """把 1 起始的列号转成 Excel 列字母（1 -> 'A'）。"""
    letters = []
    while col:
        col, rem = divmod(col - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def _patch_sheet_xml(xml_bytes: bytes, cell_values: dict) -> bytes:
    """写入数值的同时保留单元格原有样式属性。

    cell_values: {(row_no, col_no): value}，行/列均 1 起始。
    """
    root = ET.fromstring(xml_bytes)
    sheet_data = root.find(f"{{{NS}}}sheetData")
    if sheet_data is None:
        raise ValueError("worksheet has no sheetData")
    rows = {int(row.get("r")): row for row in sheet_data.findall(f"{{{NS}}}row")}
    for (row_no, col_no), value in sorted(cell_values.items()):
        row = rows.get(row_no)
        if row is None:
            row = ET.Element(f"{{{NS}}}row", {"r": str(row_no)})
            sheet_data.append(row)
            rows[row_no] = row
        ref = f"{_col_num_to_letter(col_no)}{row_no}"
        cell = next((c for c in row.findall(f"{{{NS}}}c") if c.get("r") == ref), None)
        if cell is None:
            cell = ET.SubElement(row, f"{{{NS}}}c", {"r": ref})
        cell.set("t", "n")
        for child_name in ("f", "is"):
            child = cell.find(f"{{{NS}}}{child_name}")
            if child is not None:
                cell.remove(child)
        value_node = cell.find(f"{{{NS}}}v")
        if value_node is None:
            value_node = ET.SubElement(cell, f"{{{NS}}}v")
        value_node.text = format(float(value), ".15g")
        row[:] = sorted(row, key=lambda c: _col_letter_to_num(c.get("r", "")))
    sheet_data[:] = sorted(sheet_data, key=lambda r: int(r.get("r", "0")))
    return ET.tostring(root, encoding="UTF-8", xml_declaration=True)


def _build_cell_values(plan: dict, data: ModelData) -> dict:
    """把 Q3 plan 的 x（种植面积）映射到模板单元格，按年分表。

    只写 x；b[j,t] 与 w[j,i,t,s] 属模型内部变量，不写入 Excel。
    返回: {year_str: {(row, col): area}}
    """
    values = {str(year): {} for year in data.years}
    for (j, i, year, season), area in plan["x"].items():
        plot_name = data.plot_names[j]
        col = data.tpl_crop_col.get(i)                       # 作物列
        row = (data.tpl_row_s1.get(plot_name) if season == 1  # 单季行
               else data.tpl_row_s2.get(plot_name))          # 双季行
        if col is None or row is None or str(year) not in values:
            if abs(area) > 1e-8:
                raise ValueError(f"plan cell cannot be mapped: {(j, i, year, season)}")
            continue
        key = (row, col)
        values[str(year)][key] = values[str(year)].get(key, 0.0) + float(area)
    return values


def _sheet_paths(archive: zipfile.ZipFile) -> dict:
    """sheet 名 -> 工作表 XML 在工作簿 zip 内的路径。"""
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    targets = {rel.get("Id"): rel.get("Target")
               for rel in rels.findall(f"{{{PKG_REL_NS}}}Relationship")}
    result = {}
    sheets = workbook.find(f"{{{NS}}}sheets")
    if sheets is None:
        return result
    for sheet in sheets.findall(f"{{{NS}}}sheet"):
        target = targets.get(sheet.get(f"{{{REL_NS}}}id"))
        if target:
            target = target.lstrip("/")
            if not target.startswith("xl/"):
                target = posixpath.normpath(posixpath.join("xl", target))
            result[sheet.get("name")] = target
    return result


def _sha(data: bytes) -> str:
    """原始字节的 SHA-256 十六进制摘要（用于成员级一致性比对）。"""
    return hashlib.sha256(data).hexdigest()


def _without_sheet_data(xml_bytes: bytes) -> bytes:
    """返回去掉 sheetData 的工作表 XML，用于结构（样式/格式）比对。"""
    root = ET.fromstring(xml_bytes)
    sheet_data = root.find(f"{{{NS}}}sheetData")
    if sheet_data is not None:
        root.remove(sheet_data)
    return ET.tostring(root, encoding="UTF-8")


def patch_result3(plan: dict, data: ModelData,
                  template_path: Path = paths.TEMPLATE2_PATH,
                  output_path: Path = paths.RESULT3_CANDIDATE) -> float:
    """Safely patch result3_candidate.xlsx from Q3 plan.

    Uses OOXML ZIP/XML direct manipulation (same as Q2):
      1. Copy template to candidate file
      2. Read as ZIP, modify sheet XML only for data cells
      3. Preserve all formatting/styles
      4. Read back and verify
      5. Atomic replacement via os.replace

    Returns: max readback difference (should be ~0)
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    expected = _build_cell_values(plan, data)                 # 期望写入值
    patch_values = {year: {cell: value for cell, value in cells.items()
                           if abs(value) > 1e-10}            # 只写非零格
                    for year, cells in expected.items()}
    handle = tempfile.NamedTemporaryFile(prefix=".result3-", suffix=".xlsx",
                                         dir=output_path.parent, delete=False)
    temp_path = Path(handle.name)                             # 临时候选文件
    handle.close()
    try:
        with zipfile.ZipFile(template_path, "r") as source:
            sheet_paths = _sheet_paths(source)
            target_sheets = {sheet_paths[str(year)] for year in data.years
                             if str(year) in sheet_paths}    # 需要改的表
            if len(target_sheets) != len(data.years):
                raise ValueError("template does not contain all seven year sheets")
            with zipfile.ZipFile(temp_path, "w") as target:
                for item in source.infolist():
                    payload = source.read(item.filename)
                    for year, sheet_path in sheet_paths.items():
                        if item.filename == sheet_path and year in patch_values:
                            payload = _patch_sheet_xml(payload, patch_values[year])
                            break
                    target.writestr(item, payload)

        non_target_diff = structure_diff = changed_sheets = 0
        with zipfile.ZipFile(template_path, "r") as original, \
             zipfile.ZipFile(temp_path, "r") as candidate:
            if set(original.namelist()) != set(candidate.namelist()):
                raise ValueError("OOXML member list changed")
            for name in original.namelist():
                before, after = original.read(name), candidate.read(name)
                if name in target_sheets:
                    changed_sheets += int(_sha(before) != _sha(after))
                    structure_diff += int(
                        _without_sheet_data(before) != _without_sheet_data(after))
                else:
                    non_target_diff += int(_sha(before) != _sha(after))

        workbook = openpyxl.load_workbook(temp_path, data_only=True, read_only=True)
        max_diff = 0.0                                          # 回读最大误差
        nonzero_count = 0
        for year, cells in expected.items():
            if year not in workbook.sheetnames:
                raise ValueError(f"missing output sheet {year}")
            sheet = workbook[year]
            for (row, col), expected_value in cells.items():
                actual = sheet.cell(row, col).value
                actual_value = 0.0 if actual in (None, "") else float(actual)
                max_diff = max(max_diff, abs(actual_value - expected_value))
                nonzero_count += int(abs(actual_value) > 1e-10)
        workbook.close()
        if (max_diff > 1e-4 or non_target_diff or structure_diff
                or changed_sheets != len(data.years) or nonzero_count == 0):
            raise RuntimeError(
                f"result3 OOXML audit failed: max_diff={max_diff}, "
                f"non_target_diff={non_target_diff}, structure_diff={structure_diff}, "
                f"changed_sheets={changed_sheets}, nonzero={nonzero_count}")
        os.replace(temp_path, output_path)                    # 原子替换
        return max_diff
    finally:
        temp_path.unlink(missing_ok=True)
