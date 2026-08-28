# -*- coding: utf-8 -*-
"""Focused detail dump: crop list, template header row, template plot rows,
and full 2023 statistical table.
"""
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# 相对路径：scripts/ -> q1_test/ -> practice_1/
ROOT = Path(__file__).resolve().parent.parent.parent
F1 = ROOT / "doc" / "C题" / "附件1.xlsx"
F2 = ROOT / "doc" / "C题" / "附件2.xlsx"
T1 = ROOT / "doc" / "C题" / "附件3" / "result1_1.xlsx"

def safe(v):
    if v is None:
        return ""
    s = str(v).replace("_x000d_", "").replace("\r", "").strip()
    return s

# ---- 附件1 sheet2: crop list (full) ----
wb = load_workbook(F1, data_only=True)
ws = wb["乡村种植的农作物"]
print("### 附件1 / 乡村种植的农作物 (full rows 2-42)")
print("编号 | 名称 | 类型 | 种植耕地(raw) | 说明")
for r in range(2, 43):
    code = ws.cell(r, 1).value
    name = safe(ws.cell(r, 2).value)
    typ = safe(ws.cell(r, 3).value)
    lands = safe(ws.cell(r, 4).value).replace("\n", " | ")
    note = safe(ws.cell(r, 5).value)
    if code is None:
        continue
    print(f"{code} | {name} | {typ} | {lands[:80]} | {note[:60]}")
wb.close()

# ---- 附件1 sheet1: plots (full) ----
wb = load_workbook(F1, data_only=True)
ws = wb["乡村的现有耕地"]
print("\n### 附件1 / 乡村的现有耕地 (full rows 2-55)")
print("idx | name | type | area")
for r in range(2, 56):
    name = safe(ws.cell(r, 1).value)
    typ = safe(ws.cell(r, 2).value)
    area = ws.cell(r, 3).value
    print(f"{r-1} | {name} | {typ} | {area}")
wb.close()

# ---- Template: header row (crops) and first-column plot rows ----
wb = load_workbook(T1, data_only=True)
ws = wb["2024"]
print(f"\n### result1_1 / 2024  max_col={ws.max_column} max_row={ws.max_row}")
# header row 1
hdr = []
for c in range(1, ws.max_column + 1):
    hdr.append(safe(ws.cell(1, c).value))
print("Header row1 (cols 1..end):")
print(hdr)
# plot names in col 2, rows 2..84
print("\nPlot col (B) rows 2..84:")
for r in range(2, 85):
    nm = safe(ws.cell(r, 2).value)
    a1 = safe(ws.cell(r, 1).value)
    print(f"R{r}: A='{a1}' B='{nm}'")
wb.close()

# ---- 附件2 sheet2: 2023 statistical params (full) ----
wb = load_workbook(F2, data_only=True)
ws = wb["2023年统计的相关数据"]
print(f"\n### 附件2 / 2023年统计的相关数据  max_row={ws.max_row} max_col={ws.max_column}")
print("序号 | 编号 | 名称 | 地块类型 | 季次 | 亩产量 | 成本 | 单价")
n_valid = 0
for r in range(2, ws.max_row + 1):
    seq = ws.cell(r, 1).value
    code = ws.cell(r, 2).value
    name = safe(ws.cell(r, 3).value)
    land = safe(ws.cell(r, 4).value)
    season = safe(ws.cell(r, 5).value)
    yld = ws.cell(r, 6).value
    cost = ws.cell(r, 7).value
    price = safe(ws.cell(r, 8).value)
    if code is None and name == "" and land == "":
        # note row
        print(f"NOTE R{r}: seq={seq} text={safe(ws.cell(r,2).value)[:80]}")
        continue
    n_valid += 1
    print(f"{seq} | {code} | {name} | {land} | {season} | {yld} | {cost} | {price}")
print(f"valid stat rows = {n_valid}")
wb.close()

# ---- 附件2 sheet1: 2023 planting (full) ----
wb = load_workbook(F2, data_only=True)
ws = wb["2023年的农作物种植情况"]
print(f"\n### 附件2 / 2023年的农作物种植情况  max_row={ws.max_row}")
print("row | 地块 | 编号 | 名称 | 类型 | 面积 | 季次")
n_rec = 0
prev_block = ""
for r in range(2, ws.max_row + 1):
    blk = safe(ws.cell(r, 1).value)
    if blk == "":
        blk = prev_block
    else:
        prev_block = blk
    code = ws.cell(r, 2).value
    name = safe(ws.cell(r, 3).value)
    typ = safe(ws.cell(r, 4).value)
    area = ws.cell(r, 5).value
    season = safe(ws.cell(r, 6).value)
    if code is None and name == "":
        continue
    n_rec += 1
    print(f"R{r} | {blk} | {code} | {name} | {typ} | {area} | {season}")
print(f"valid planting records = {n_rec}")
wb.close()
