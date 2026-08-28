# -*- coding: utf-8 -*-
"""One-off data-structure exploration script.

Dumps each worksheet's shape, header, first/last few rows, and merged ranges
for 附件1/附件2 and result1_1/result1_2 templates.
"""
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# 相对路径：scripts/ -> q1_test/ -> practice_1/
ROOT = Path(__file__).resolve().parent.parent.parent
F1 = ROOT / "doc" / "C题" / "附件1.xlsx"
F2 = ROOT / "doc" / "C题" / "附件2.xlsx"
T1 = ROOT / "doc" / "C题" / "附件3" / "result1_1.xlsx"
T2 = ROOT / "doc" / "C题" / "附件3" / "result1_2.xlsx"


def dump_xlsx(path, max_rows=6):
    print("=" * 80)
    print("FILE:", path.name)
    wb = load_workbook(path, data_only=True)
    print("Sheets:", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print("-" * 70)
        print(f"  Sheet: {sn!r}  dims={ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}")
        merges = list(ws.merged_cells.ranges)
        print(f"  merged_ranges({len(merges)}):", [str(m) for m in merges[:8]])
        for r in range(1, min(ws.max_row, max_rows) + 1):
            row_vals = []
            for c in range(1, min(ws.max_column, 20) + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    row_vals.append("·")
                else:
                    s = str(v).replace("\n", "\\n").replace("\r", "\\r")
                    if len(s) > 30:
                        s = s[:27] + "..."
                    row_vals.append(s)
            print(f"   R{r:>3}: " + " | ".join(row_vals))
        if ws.max_row > max_rows:
            for r in range(max(ws.max_row - 2, max_rows + 1), ws.max_row + 1):
                row_vals = []
                for c in range(1, min(ws.max_column, 20) + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is None:
                        row_vals.append("·")
                    else:
                        s = str(v).replace("\n", "\\n").replace("\r", "\\r")
                        if len(s) > 30:
                            s = s[:27] + "..."
                        row_vals.append(s)
                print(f"   R{r:>3}: " + " | ".join(row_vals))
    wb.close()


for p in [F1, F2, T1, T2]:
    try:
        dump_xlsx(p)
    except Exception as e:
        print("ERROR for", p, "->", repr(e))
