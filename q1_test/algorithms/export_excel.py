# -*- coding: utf-8 -*-
"""Excel output (AGENT.md section 9).

  export_result_workbook(sol, data, template_path, output_path)
    - copy the official template (never overwrite the original)
    - write planted area x into the correct (year-sheet, plot-row, crop-col)
    - preserve sheet names, layout, styles, merged cells, crop order
    - zero-fill the data region for a consistent empty-value convention
  reread_audit(workbook_path, sol, data) -> max |cell - x|  (tolerance 1e-4)
"""
from __future__ import annotations
import shutil
from pathlib import Path
from openpyxl import load_workbook

from .preprocess import ModelData


def _data_rows(data: ModelData):
    """Yield (plot_name, season, row_index) for every template data cell."""
    for name, r in data.tpl_row_s1.items():
        yield name, 1, r
    for name, r in data.tpl_row_s2.items():
        yield name, 2, r


def export_result_workbook(sol: dict, data: ModelData,
                           template_path: Path, output_path: Path) -> Path:
    """Copy template, fill area values, save. Returns output path."""
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    # build x lookup: (plot_name, crop_code, year, season) -> area
    x_lookup = {}
    if sol.get("x") is not None:
        for _, r in sol["x"].iterrows():
            x_lookup[(r["plot"], int(r["crop_code"]), int(r["year"]),
                      int(r["season"]))] = float(r["area"])

    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        try:
            year = int(sheet_name)
        except ValueError:
            continue            # skip non-year sheets
        if year not in data.years:
            continue
        ws = wb[sheet_name]
        # zero-fill data region then write nonzero values
        for plot_name, season, row in _data_rows(data):
            for code in data.crop_codes:
                col = data.tpl_crop_col[code]
                # default 0 for consistency
                area = x_lookup.get((plot_name, code, year, season), 0.0)
                ws.cell(row=row, column=col).value = area
    wb.save(output_path)
    wb.close()
    return output_path


def reread_audit(workbook_path: Path, sol: dict, data: ModelData) -> float:
    """Reload workbook, compare every data cell to in-memory x. Max diff."""
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    x_lookup = {}
    if sol.get("x") is not None:
        for _, r in sol["x"].iterrows():
            x_lookup[(r["plot"], int(r["crop_code"]), int(r["year"]),
                      int(r["season"]))] = float(r["area"])
    max_diff = 0.0
    for sheet_name in wb.sheetnames:
        try:
            year = int(sheet_name)
        except ValueError:
            continue
        if year not in data.years:
            continue
        ws = wb[sheet_name]
        for plot_name, season, row in _data_rows(data):
            for code in data.crop_codes:
                col = data.tpl_crop_col[code]
                cell = ws.cell(row=row, column=col).value
                expected = x_lookup.get((plot_name, code, year, season), 0.0)
                if isinstance(cell, bool) or not isinstance(cell, (int, float)):
                    raise ValueError(
                        f"Non-numeric result cell {sheet_name}!"
                        f"{ws.cell(row=row, column=col).coordinate}: {cell!r}")
                got = float(cell)
                d = abs(got - expected)
                if d > max_diff:
                    max_diff = d
    wb.close()
    return max_diff
