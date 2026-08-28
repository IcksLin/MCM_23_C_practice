# -*- coding: utf-8 -*-
"""Read-only Excel ingestion.

Reads 附件1 (plots & crops), 附件2 (2023 planting & statistics) and the
result1_1 template structure into a RawData container. Merged-cell values
in the planting sheet are forward-filled here so downstream cleaning works
on a tidy table; all string stripping, type conversion and parameter
inheritance belong to preprocess.py.
"""
from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
import hashlib
import pandas as pd
from openpyxl import load_workbook

from . import paths


def _safe(v):
    if v is None:
        return ""
    return str(v)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest().upper()


def verify_inputs() -> dict:
    """Return {path: (sha, ok)} for every frozen input."""
    out = {}
    for p, exp in paths.EXPECTED_SHA.items():
        sha = sha256_file(p)
        out[str(p)] = (sha, sha == exp, exp)
    return out


@dataclass
class RawData:
    plots: pd.DataFrame            # name, type, area
    crops: pd.DataFrame            # code, name, type, lands_raw, note
    planting_2023: pd.DataFrame    # plot, crop_code, crop_name, crop_type, area, season
    stats_2023: pd.DataFrame       # seq, crop_code, crop_name, land_type, season, yield, cost, price_raw
    template_crops: list           # crop names in template column order (len 41)
    template_plot_s1: list         # plot names in 第一季 rows (len 54)
    template_plot_s2: list         # plot names in 第二季 rows (len 28)
    f1_path: Path
    f2_path: Path
    template1_path: Path
    f1_sha: str
    f2_sha: str
    template1_sha: str


def _read_plots(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb["乡村的现有耕地"]
    rows = []
    for r in range(2, ws.max_row + 1):
        name = _safe(ws.cell(r, 1).value)
        typ = _safe(ws.cell(r, 2).value)
        area = ws.cell(r, 3).value
        if name == "" and typ == "":
            continue
        rows.append({"name": name, "type": typ, "area": area})
    wb.close()
    return pd.DataFrame(rows)


def _is_int_like(v) -> bool:
    if isinstance(v, (int,)) and not isinstance(v, bool):
        return True
    if isinstance(v, float) and v.is_integer():
        return True
    if isinstance(v, str):
        try:
            int(v.strip())
            return True
        except ValueError:
            return False
    return False


def _read_crops(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb["乡村种植的农作物"]
    rows = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not _is_int_like(code):
            continue                       # skip 注： / blank rows
        name = _safe(ws.cell(r, 2).value)
        typ = _safe(ws.cell(r, 3).value)
        lands = _safe(ws.cell(r, 4).value)
        note = _safe(ws.cell(r, 5).value)
        rows.append({
            "code": int(code),
            "name": name,
            "type": typ,
            "lands_raw": lands,
            "note": note,
        })
    wb.close()
    return pd.DataFrame(rows)


def _read_planting(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb["2023年的农作物种植情况"]
    rows = []
    prev_block = ""
    for r in range(2, ws.max_row + 1):
        blk = _safe(ws.cell(r, 1).value)
        if blk == "":
            blk = prev_block
        else:
            prev_block = blk
        code = ws.cell(r, 2).value
        name = _safe(ws.cell(r, 3).value)
        typ = _safe(ws.cell(r, 4).value)
        area = ws.cell(r, 5).value
        season = _safe(ws.cell(r, 6).value)
        if code is None and name == "":
            continue
        rows.append({
            "plot": blk,
            "crop_code": code,
            "crop_name": name,
            "crop_type": typ,
            "area": area,
            "season": season,
        })
    wb.close()
    return pd.DataFrame(rows)


def _read_stats(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb["2023年统计的相关数据"]
    rows = []
    for r in range(2, ws.max_row + 1):
        seq = ws.cell(r, 1).value
        code = ws.cell(r, 2).value
        # note rows: seq == "注：" or crop_code not numeric
        if isinstance(seq, str) and "注" in seq:
            continue
        if not _is_int_like(code):
            continue
        name = _safe(ws.cell(r, 3).value)
        land = _safe(ws.cell(r, 4).value)
        season = _safe(ws.cell(r, 5).value)
        yld = ws.cell(r, 6).value
        cost = ws.cell(r, 7).value
        price = _safe(ws.cell(r, 8).value)
        rows.append({
            "seq": seq,
            "crop_code": int(code),
            "crop_name": name,
            "land_type": land,
            "season": season,
            "yield": yld,
            "cost": cost,
            "price_raw": price,
        })
    wb.close()
    return pd.DataFrame(rows)


def _read_template(path: Path) -> tuple[list, list, list]:
    wb = load_workbook(path, data_only=True)
    ws = wb["2024"]
    crops = []
    for c in range(3, ws.max_column + 1):
        v = _safe(ws.cell(1, c).value)
        if v != "":
            crops.append(v)
    # season-1 plot rows 2..55
    s1 = [_safe(ws.cell(r, 2).value) for r in range(2, 56)]
    # season-2 plot rows 56..83
    s2 = [_safe(ws.cell(r, 2).value) for r in range(56, 84)]
    wb.close()
    return crops, s1, s2


def load_inputs() -> RawData:
    """Read all inputs (read-only). Raises on hash mismatch."""
    ver = verify_inputs()
    bad = [p for p, (_, ok, _) in ver.items() if not ok]
    if bad:
        raise RuntimeError(f"Input hash mismatch, refusing to proceed: {bad}")
    plots = _read_plots(paths.F1_PATH)
    crops = _read_crops(paths.F1_PATH)
    planting = _read_planting(paths.F2_PATH)
    stats = _read_stats(paths.F2_PATH)
    template_crops, tpl_s1, tpl_s2 = _read_template(paths.TEMPLATE1_PATH)
    return RawData(
        plots=plots,
        crops=crops,
        planting_2023=planting,
        stats_2023=stats,
        template_crops=template_crops,
        template_plot_s1=tpl_s1,
        template_plot_s2=tpl_s2,
        f1_path=paths.F1_PATH,
        f2_path=paths.F2_PATH,
        template1_path=paths.TEMPLATE1_PATH,
        f1_sha=ver[str(paths.F1_PATH)][0],
        f2_sha=ver[str(paths.F2_PATH)][0],
        template1_sha=ver[str(paths.TEMPLATE1_PATH)][0],
    )
