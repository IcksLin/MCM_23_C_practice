# -*- coding: utf-8 -*-
"""只读验证三问冻结工作簿、镜像输入和Q1真实复现结果。"""
from __future__ import annotations
import argparse, subprocess, sys
from pathlib import Path
import openpyxl, pandas as pd

ROOT = Path(__file__).resolve().parent.parent


def numeric_map(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    out[(ws.title, cell.row, cell.column)] = float(cell.value)
    wb.close()
    return out


def verify_q1() -> None:
    for name in ("result1_1.xlsx", "result1_2.xlsx"):
        ref = numeric_map(ROOT / "doc/results/q1" / name)
        rep = numeric_map(ROOT / "doc/results/q1/reproduced" / name)
        keys = set(ref) | set(rep)
        diff = max((abs(ref.get(k, 0.0) - rep.get(k, 0.0)) for k in keys), default=0.0)
        nonzero = sum(abs(v) > 1e-10 for v in ref.values())
        if diff > 1e-4 or nonzero == 0:
            raise SystemExit(f"Q1校验失败 {name}: diff={diff}, nonzero={nonzero}")
        print(f"Q1 {name}: nonzero={nonzero}, reproduced_max_diff={diff:.3e}")


def verify_plan(problem: str) -> None:
    sys.path.insert(0, str(ROOT / f"{problem}_test"))
    from algorithms.io_data import load_inputs
    from algorithms.preprocess import preprocess
    from algorithms.export_ooxml import _build_cell_values
    data = preprocess(load_inputs())
    if problem == "q2":
        csv_path = ROOT / "doc/results/q2/selected_plan.csv"
        book_path = ROOT / "doc/results/q2/result2.xlsx"
    else:
        csv_path = ROOT / "doc/results/q3/selected_plan_q3.csv"
        book_path = ROOT / "doc/results/q3/result3.xlsx"
    frame = pd.read_csv(csv_path)
    x = {(data.plot_idx[str(r.plot)], int(r.crop_code), int(r.year), int(r.season)):
         float(r.area) for r in frame.itertuples(index=False)}
    expected = _build_cell_values({"x": x}, data)
    wb = openpyxl.load_workbook(book_path, data_only=True, read_only=True)
    diffs = [abs(float(wb[y].cell(row, col).value or 0.0) - value)
             for y, cells in expected.items() for (row, col), value in cells.items()]
    wb.close()
    diff = max(diffs, default=float("inf"))
    if diff > 1e-4 or not diffs:
        raise SystemExit(f"{problem.upper()}工作簿校验失败: cells={len(diffs)}, diff={diff}")
    print(f"{problem.upper()}: mapped_cells={len(diffs)}, max_diff={diff:.3e}")


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--problem", choices=["q1", "q2", "q3"])
    args = p.parse_args()
    if args.problem == "q1": verify_q1()
    elif args.problem in ("q2", "q3"): verify_plan(args.problem)
    else:
        for problem in ("q1", "q2", "q3"):
            result = subprocess.run([sys.executable, str(Path(__file__).resolve()),
                                     "--problem", problem], cwd=ROOT)
            if result.returncode:
                return result.returncode
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
